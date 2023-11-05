from reportlab.lib.pagesizes import landscape, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph
from reportlab.lib.enums import TA_CENTER
from openpyxl.styles.alignment import Alignment
from openpyxl import Workbook

def generate_pdf_report(data):
    pdf_report = SimpleDocTemplate("report.pdf", pagesize=landscape(letter))
    elements = []


    style = getSampleStyleSheet()
    style = style["BodyText"]
    style.alignment = TA_CENTER
    style.wordWrap = 'CJK'


    table_data = [["Order ID", "Nombre", "Email", "Dirección", "Productos", "Valor Total", "Fecha"]]
    for order in data:
        table_data.append([
            order.id,
            Paragraph(order.name, style),
            Paragraph(order.email, style),
            Paragraph(f"{order.city},\n{order.address}", style),
            Paragraph(order.product, style),
            order.total_price,
            order.created.strftime("%Y-%m-%d %H:%M:%S"),
        ])


    col_widths = [50, 100, 100, 150, 100, 50, 100]

    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 12))

    # Build the PDF report
    pdf_report.build(elements)



def generate_excel_report(data, created_field_name):
    workbook = Workbook()
    sheet = workbook.active

    # Add data to the Excel report
    sheet.append(["Order ID", "Nombre", "Email", "Dirección", "Productos", "Valor Total", "Fecha"])
    for order in data:
        created_without_timezone = getattr(order, created_field_name).replace(tzinfo=None)
        row = [
            order.id,
            order.name,
            order.email,
            f"{order.city}, {order.address}",
            order.product,
            order.total_price,
            created_without_timezone,
        ]
        sheet.append(row)


    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    workbook.save("report.xlsx")
